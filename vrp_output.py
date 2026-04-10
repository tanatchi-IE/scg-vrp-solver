# =====================================================
# FILE: vrp_output.py
# OUTPUT MODULE — Excel Export, Charts, Map
# Version: 3.1 (Based on Part D + matplotlib + pickup path)
# =====================================================
import io
import math
import zipfile
from typing import List, Dict, Tuple, Optional
from collections import defaultdict
from datetime import datetime
import pandas as pd
import numpy as np

# =====================================================
# SECTION 1: CONSTANTS & HELPERS
# =====================================================
ROUTE_COLORS = [
    '#e6194b', '#3cb44b', '#4363d8', '#f58231', '#911eb4',
    '#42d4f4', '#f032e6', '#bfef45', '#fabed4', '#469990',
    '#dcbeff', '#9A6324', '#fffac8', '#800000', '#aaffc3',
    '#808000', '#ffd8b1', '#000075', '#a9a9a9', '#000000',
]

DEPOT_ICON_COLORS = {
    '329': 'red', '0329': 'red',
    '03T7': 'blue', '3T7': 'blue',
}

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from vrp_core import (
        DEFAULT_PRIORITY_CONFIG, minutes_to_time_str,
        get_priority_name, get_order_deadline,
        normalize_depot_id,
    )
except ImportError:
    DEFAULT_PRIORITY_CONFIG = {
        1: {'name': 'Critical', 'deadline': 600, 'hard': True},
        2: {'name': 'Urgent', 'deadline': 720, 'hard': True},
        3: {'name': 'Afternoon', 'deadline': 900, 'hard': False},
        4: {'name': 'Normal', 'deadline': None, 'hard': False},
        5: {'name': 'Flexible', 'deadline': 1080, 'hard': False},
    }
    def minutes_to_time_str(minutes) -> str:
        if minutes is None or minutes == '':
            return "TW Close"
        try:
            minutes = int(minutes)
            return f"{minutes // 60:02d}:{minutes % 60:02d}"
        except (ValueError, TypeError):
            return str(minutes)
    def get_priority_name(priority, config=None):
        if config is None: config = DEFAULT_PRIORITY_CONFIG
        return config.get(priority, {}).get('name', f'Priority {priority}')
    def get_order_deadline(order, config=None):
        if config is None: config = DEFAULT_PRIORITY_CONFIG
        p = getattr(order, 'priority', 4)
        c = config.get(p, {})
        if c and c.get('deadline') is not None:
            return c['deadline']
        return getattr(order, 'time_close', 1020)


def _safe_get(obj, attr, default=0):
    val = getattr(obj, attr, default)
    return val if val is not None else default


def generate_google_maps_link(depot_lat, depot_lng, stops):
    """สร้าง Google Maps Direction Link"""
    base_url = "https://www.google.com/maps/dir"
    coords = [f"{depot_lat},{depot_lng}"]
    for stop in stops:
        if hasattr(stop, 'order'):
            coords.append(f"{stop.order.lat},{stop.order.lng}")
        elif hasattr(stop, 'lat'):
            coords.append(f"{stop.lat},{stop.lng}")
    return base_url + "/" + "/".join(coords)


# =====================================================
# SECTION 2: SUMMARY REPORT GENERATOR
# =====================================================
def generate_summary_report(solution, data) -> Dict:
    """สร้าง Summary Report จาก solution — ใช้ใน app.py"""
    summary = solution.get_summary()
    dbd = solution.display_breakdown
    bd = solution.breakdown

    # ── Depots info ──
    depots = {}
    if data is not None:
        if hasattr(data, 'depots'):
            depots = data.depots
        elif isinstance(data, dict):
            depots = data.get('depots', {})

    priority_config = DEFAULT_PRIORITY_CONFIG
    if data is not None:
        if hasattr(data, 'priority_config'):
            priority_config = data.priority_config
        elif isinstance(data, dict):
            priority_config = data.get('priority_config', DEFAULT_PRIORITY_CONFIG)

    # ── Overview ──
    overview = {
        'total_cost': summary['total_cost'],
        'num_vehicles': summary['num_vehicles'],
        'total_orders': summary['total_orders'],
        'assigned_orders': summary['assigned_orders'],
        'unassigned_orders': summary['unassigned_orders'],
        'assignment_rate': (summary['assigned_orders'] / max(1, summary['total_orders'])) * 100,
        'total_distance': summary['total_distance'],
        'pickup_distance': summary['pickup_distance'],
        'delivery_distance': summary['delivery_distance'],
        'num_late': summary['num_late'],
        'num_priority_miss': summary['num_priority_miss'],
        'overtime_hours': summary['overtime_hours'],
        'extra_drops': summary['extra_drops'],
        'multi_depot_routes': summary.get('multi_depot_routes', 0),
        'single_depot_routes': summary.get('single_depot_routes', 0),
    }

    # ── Cost Breakdown ──
    cost_breakdown = {
        'fixed_cost': dbd.get('fixed_cost', 0),
        'variable_cost': dbd.get('variable_cost', 0),
        'unassigned_cost': dbd.get('unassigned_cost', 0),
        'late_cost': dbd.get('late_cost', 0),
        'overtime_cost': dbd.get('overtime_cost', 0),
        'extra_drop_cost': dbd.get('extra_drop_cost', 0),
        'extra_drop_details': dbd.get('extra_drop_details', []),
    }

    # ── Priority Summary ──
    priority_summary = []
    pbd = dbd.get('priority_breakdown', {})
    if hasattr(pbd, 'items'):
        for p in sorted(pbd.keys()):
            pdata = pbd[p]
            priority_summary.append({
                'priority': p,
                'name': get_priority_name(p, priority_config),
                'assigned': pdata.get('assigned', 0),
                'missed': pdata.get('missed', 0),
                'status': '✅' if pdata.get('missed', 0) == 0 else f"⚠️ {pdata['missed']} missed"
            })

    # ── Route Summaries ──
    route_summaries = []
    active_routes = [r for r in solution.routes if not r.is_empty()]
    active_routes_sorted = sorted(active_routes, key=lambda r: r.vehicle.vehicle_id)

    for i, route in enumerate(active_routes_sorted, 1):
        v = route.vehicle
        stops = route.stops
        first_stop = stops[0] if stops else None
        last_stop = stops[-1] if stops else None

        depot_names = []
        if hasattr(route, 'pickup_sequence') and route.pickup_sequence:
            for d in route.pickup_sequence:
                depot_names.append(getattr(d, 'name', str(d)))

        first_arrival = _safe_get(first_stop, 'arrival_time', 0) if first_stop else 0
        last_departure = _safe_get(last_stop, 'departure_time', 0) if last_stop else 0
        late_count = sum(1 for s in stops if _safe_get(s, 'is_late', False))
        weight_util = (route.total_weight / v.max_weight_kg * 100) if v.max_weight_kg > 0 else 0
        volume_util = (route.total_volume / v.max_volume_cbm * 100) if v.max_volume_cbm > 0 else 0

        start_time = getattr(v, 'start_time', 480)
        if hasattr(start_time, 'hour'):
            start_time = start_time.hour * 60 + start_time.minute
        working_hours = (last_departure - start_time) / 60 if last_departure > 0 else 0

        capacity_drop = getattr(v, 'capacity_drop', 7)

        route_summaries.append({
            'route_number': i,
            'vehicle_id': v.vehicle_id,
            'vehicle_type': getattr(v, 'vehicle_type', '4W'),
            'driver_name': getattr(v, 'driver_name', 'N/A'),
            'num_stops': route.num_stops,
            'capacity_drop': capacity_drop,
            'depots': depot_names,
            'num_depots': len(set(s.order.plant for s in stops)),
            'total_weight': route.total_weight,
            'max_weight': v.max_weight_kg,
            'weight_util': weight_util,
            'total_volume': route.total_volume,
            'max_volume': v.max_volume_cbm,
            'volume_util': volume_util,
            'total_distance': _safe_get(route, 'total_distance', 0),
            'pickup_distance': _safe_get(route, 'pickup_distance', 0),
            'delivery_distance': _safe_get(route, 'delivery_distance', 0),
            'first_arrival': minutes_to_time_str(first_arrival),
            'last_departure': minutes_to_time_str(last_departure),
            'working_hours': working_hours,
            'late_count': late_count,
            'stops_detail': _build_stops_detail(stops, priority_config),
        })

    # ── Unassigned detail (with reason) ──
    unassigned_detail = []
    for order in solution.unassigned:
        unassigned_detail.append({
            'customer_name': order.customer_name,
            'plant': order.plant,
            'priority': getattr(order, 'priority', 4),
            'weight_kg': order.weight_kg,
            'volume_cbm': order.volume_cbm,
            'reason': getattr(order, 'unassign_reason', '') or 'ไม่ทราบสาเหตุ',
        })

    # ── Warnings ──
    warnings = []
    if overview['unassigned_orders'] > 0:
        names = [o.customer_name for o in solution.unassigned[:3]]
        suffix = f" และอีก {len(solution.unassigned) - 3} ราย" if len(solution.unassigned) > 3 else ""
        warnings.append(f"⚠️ จัดไม่ได้ {overview['unassigned_orders']} ออเดอร์: {', '.join(names)}{suffix}")
    if overview['num_late'] > 0:
        warnings.append(f"⚠️ ส่งสาย {overview['num_late']} จุด")
    if overview['num_priority_miss'] > 0:
        warnings.append(f"🚨 Priority miss {overview['num_priority_miss']} ออเดอร์")
    if overview['overtime_hours'] > 0:
        warnings.append(f"⏰ OT รวม {overview['overtime_hours']:.1f} ชม.")

    return {
        'overview': overview,
        'cost_breakdown': cost_breakdown,
        'priority_summary': priority_summary,
        'route_summaries': route_summaries,
        'unassigned_detail': unassigned_detail,
        'warnings': warnings,
    }


def _build_stops_detail(stops, priority_config=None) -> List[Dict]:
    if priority_config is None:
        priority_config = DEFAULT_PRIORITY_CONFIG
    details = []
    for idx, stop in enumerate(stops):
        order = stop.order
        priority = getattr(order, 'priority', 4)
        deadline = get_order_deadline(order, priority_config)
        details.append({
            'stop_number': idx + 1,
            'customer_name': order.customer_name,
            'dn_numbers': getattr(order, 'dn_numbers', []),
            'plant': order.plant,
            'priority': priority,
            'priority_name': get_priority_name(priority, priority_config),
            'deadline': minutes_to_time_str(deadline),
            'weight_kg': order.weight_kg,
            'volume_cbm': order.volume_cbm,
            'arrival_time': minutes_to_time_str(_safe_get(stop, 'arrival_time', 0)),
            'departure_time': minutes_to_time_str(_safe_get(stop, 'departure_time', 0)),
            'time_window': f"{minutes_to_time_str(getattr(order, 'time_open', 480))}-{minutes_to_time_str(getattr(order, 'time_close', 1020))}",
            'distance_from_prev': round(_safe_get(stop, 'distance_from_prev', 0), 1),
            'cumulative_distance': round(_safe_get(stop, 'cumulative_distance', 0), 1),
            'wait_time': _safe_get(stop, 'wait_time', 0),
            'is_late': _safe_get(stop, 'is_late', False),
            'zone': getattr(order, 'zone', ''),
            'district': getattr(order, 'district', ''),
            'province': getattr(order, 'province', ''),
            'lat': order.lat,
            'lng': order.lng,
            'ship_to_code': getattr(order, 'ship_to_code', ''),
        })
    return details


# =====================================================
# SECTION 3: ROUTE DETAIL EXCEL (Part D — Lean Format)
# =====================================================
def create_lean_route_df(routes, depots, priority_config=None):
    """สร้าง DataFrame แบบ Lean — เหมือน Part D เดิมเป๊ะ"""
    if priority_config is None:
        priority_config = DEFAULT_PRIORITY_CONFIG

    output_rows = []

    default_depot = None
    for d in depots.values():
        if getattr(d, 'is_default_end', False):
            default_depot = d
            break
    if default_depot is None and depots:
        default_depot = list(depots.values())[0]

    depot_lat = default_depot.lat if default_depot else 14.0257
    depot_lng = default_depot.lng if default_depot else 100.6141

    sorted_routes = sorted(
        [r for r in routes if not r.is_empty()],
        key=lambda r: r.vehicle.vehicle_id
    )

    for route_idx, route in enumerate(sorted_routes, 1):
        vehicle = route.vehicle

        # Depot names
        if hasattr(route, 'pickup_sequence') and route.pickup_sequence:
            depot_names = " → ".join([d.name for d in route.pickup_sequence])
        elif hasattr(route, 'orders_by_depot') and route.orders_by_depot:
            depot_ids = list(route.orders_by_depot.keys())
            depot_names = " → ".join([
                depots[d].name if d in depots else d for d in depot_ids
            ])
        else:
            depot_names = default_depot.name if default_depot else "DC"

        route_weight = 0
        route_volume = 0
        late_count = 0
        start_time = getattr(vehicle, 'start_time', 480)
        if hasattr(start_time, 'hour'):
            start_time = start_time.hour * 60 + start_time.minute

        # ── Delivery Stops ──
        for i, stop in enumerate(route.stops):
            order = stop.order
            time_close = getattr(order, 'time_close', 1020)
            arrival = getattr(stop, 'arrival_time', 0)
            is_late = arrival > time_close

            deadline = get_order_deadline(order, priority_config)
            if arrival > deadline:
                is_late = True
            if is_late:
                late_count += 1

            time_open = getattr(order, 'time_open', 480)
            time_window = f"{minutes_to_time_str(time_open)}-{minutes_to_time_str(time_close)}"
            route_weight += order.weight_kg
            route_volume += order.volume_cbm

            priority = getattr(order, 'priority', 4)
            priority_name = get_priority_name(priority, priority_config)
            deadline_str = minutes_to_time_str(deadline)

            output_rows.append({
                'Route_No': route_idx,
                'Vehicle_ID': vehicle.vehicle_id,
                'Vehicle_Type': getattr(vehicle, 'vehicle_type', '4W'),
                'Driver': getattr(vehicle, 'driver_name', 'N/A'),
                'Depot_Names': depot_names,
                'Stop_No': i + 1,
                'Customer_Name': order.customer_name,
                'District': getattr(order, 'district', ''),
                'Province': getattr(order, 'province', ''),
                'Zone': getattr(order, 'zone', ''),
                'Plant': getattr(order, 'plant', ''),
                'Priority': priority,
                'Priority_Name': priority_name,
                'Deadline': deadline_str,
                'Weight_kg': round(order.weight_kg, 2),
                'Volume_CBM': round(order.volume_cbm, 5),
                'Distance_From_Prev_km': round(getattr(stop, 'distance_from_prev', 0), 1),
                'Cumulative_Distance_km': round(getattr(stop, 'cumulative_distance', 0), 1),
                'Arrival_Time': minutes_to_time_str(arrival),
                'Departure_Time': minutes_to_time_str(getattr(stop, 'departure_time', 0)),
                'Time_Window': time_window,
                'Latitude': order.lat,
                'Longitude': order.lng,
                '_is_late': is_late,
                '_row_type': 'data'
            })

        # ── Summary Row ──
        max_weight = vehicle.max_weight_kg
        max_volume = vehicle.max_volume_cbm
        weight_pct = (route_weight / max_weight * 100) if max_weight > 0 else 0
        volume_pct = (route_volume / max_volume * 100) if max_volume > 0 else 0

        if route.stops:
            end_time = route.stops[-1].departure_time
            working_hours = (end_time - start_time) / 60
        else:
            working_hours = 0

        num_drops = len(route.stops)
        capacity_drop = getattr(vehicle, 'capacity_drop', 7)
        maps_link = generate_google_maps_link(depot_lat, depot_lng, route.stops)

        output_rows.append({
            'Route_No': '', 'Vehicle_ID': '', 'Vehicle_Type': '',
            'Driver': '', 'Depot_Names': '', 'Stop_No': '',
            'Customer_Name': '', 'District': '', 'Province': '',
            'Zone': '', 'Plant': '', 'Priority': '',
            'Priority_Name': '', 'Deadline': '',
            'Weight_kg': f'{route_weight:.0f}/{max_weight:.0f} ({weight_pct:.0f}%)',
            'Volume_CBM': f'{route_volume:.2f}/{max_volume:.1f} ({volume_pct:.0f}%)',
            'Distance_From_Prev_km': round(route.total_distance, 1),
            'Cumulative_Distance_km': '',
            'Arrival_Time': '',
            'Departure_Time': f'{working_hours:.1f} hrs',
            'Time_Window': f'{num_drops}/{capacity_drop} drops',
            'Latitude': '',
            'Longitude': maps_link,
            '_is_late': False,
            '_row_type': 'summary'
        })

    return pd.DataFrame(output_rows)


# =====================================================
# SECTION 4: EXPORT ROUTE_DETAIL.XLSX (3 sheets)
# =====================================================
def export_route_detail_excel(solution, depots, priority_config=None):
    """
    Export Route_Detail.xlsx — 3 sheets:
    1. Route_Details (Lean format + Priority + Google Maps)
    2. Route_Summary
    3. Unassigned
    Returns: BytesIO
    """
    if priority_config is None:
        priority_config = DEFAULT_PRIORITY_CONFIG

    routes = [r for r in solution.routes if not r.is_empty()]
    routes = sorted(routes, key=lambda r: r.vehicle.vehicle_id)
    df = create_lean_route_df(routes, depots, priority_config)

    output = io.BytesIO()

    if not OPENPYXL_AVAILABLE:
        display_cols = [c for c in df.columns if not c.startswith('_')]
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df[display_cols].to_excel(writer, sheet_name='Route_Details', index=False)
        output.seek(0)
        return output

    wb = Workbook()
    ws = wb.active
    ws.title = "Route_Details"

    # ── Styles ──
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    late_font = Font(bold=True, color="FF0000")
    summary_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    summary_font = Font(bold=True)
    link_font = Font(bold=True, color="0563C1", underline="single")

    priority_fills = {
        1: PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
        2: PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
        3: PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
        4: PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        5: PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    }

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    display_columns = [
        'Route_No', 'Vehicle_ID', 'Vehicle_Type', 'Driver', 'Depot_Names',
        'Stop_No', 'Customer_Name', 'District', 'Province', 'Zone', 'Plant',
        'Priority', 'Priority_Name', 'Deadline',
        'Weight_kg', 'Volume_CBM', 'Distance_From_Prev_km', 'Cumulative_Distance_km',
        'Arrival_Time', 'Departure_Time', 'Time_Window', 'Latitude', 'Longitude'
    ]

    # ── Write Header ──
    for col_idx, col_name in enumerate(display_columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── Write Data ──
    for row_idx, (_, row_data) in enumerate(df.iterrows()):
        excel_row = row_idx + 2
        is_late = row_data.get('_is_late', False)
        row_type = row_data.get('_row_type', 'data')
        priority = row_data.get('Priority', 4)

        for col_idx, col_name in enumerate(display_columns, 1):
            value = row_data.get(col_name, '')
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = thin_border

            if row_type == 'summary':
                cell.fill = summary_fill
                cell.font = summary_font
                cell.alignment = Alignment(horizontal="center")
                if col_name == 'Longitude' and str(value).startswith('http'):
                    cell.hyperlink = value
                    cell.font = link_font
                    cell.value = "Open Map"
            elif row_type == 'data':
                if col_name == 'Arrival_Time' and is_late:
                    cell.font = late_font
                if col_name == 'Priority' and isinstance(priority, int) and priority in priority_fills:
                    cell.fill = priority_fills[priority]
                if col_name in ['Route_No', 'Stop_No', 'Priority', 'Priority_Name',
                               'Deadline', 'Weight_kg', 'Volume_CBM',
                               'Distance_From_Prev_km', 'Cumulative_Distance_km',
                               'Arrival_Time', 'Departure_Time', 'Time_Window']:
                    cell.alignment = Alignment(horizontal="center")

    # ── Column Widths ──
    widths = {
        'A': 8, 'B': 10, 'C': 10, 'D': 12, 'E': 18, 'F': 6,
        'G': 40, 'H': 20, 'I': 15, 'J': 8, 'K': 6,
        'L': 8, 'M': 12, 'N': 10,
        'O': 16, 'P': 16, 'Q': 15, 'R': 18,
        'S': 10, 'T': 12, 'U': 12, 'V': 10, 'W': 12
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = 'A2'

    # ── Sheet 2: Route_Summary ──
    ws_summary = wb.create_sheet("Route_Summary")
    _add_route_summary_sheet(ws_summary, routes, depots, solution.unassigned,
                             header_fill, header_font, thin_border, priority_config)

    # ── Sheet 3: Unassigned ──
    if solution.unassigned:
        ws_unassigned = wb.create_sheet("Unassigned")
        _add_unassigned_sheet(ws_unassigned, solution.unassigned,
                              header_fill, header_font, thin_border, priority_config)

    # Save to BytesIO
    wb.save(output)
    output.seek(0)
    return output


def _add_route_summary_sheet(ws, routes, depots, unassigned,
                             header_fill, header_font, border, priority_config):
    """Sheet Route_Summary — เหมือน Part D เดิม"""
    default_depot = None
    for d in depots.values():
        if getattr(d, 'is_default_end', False):
            default_depot = d
            break
    if default_depot is None and depots:
        default_depot = list(depots.values())[0]

    depot_lat = default_depot.lat if default_depot else 14.0257
    depot_lng = default_depot.lng if default_depot else 100.6141

    headers = ['Route', 'Vehicle', 'Driver', 'Type', 'Depots', 'Drops',
               'Weight', 'Weight%', 'Volume', 'Vol%', 'Distance',
               'Hours', 'Late', 'Cost', 'Map']

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for row_idx, route in enumerate(routes, 2):
        v = route.vehicle
        weight_pct = (route.total_weight / v.max_weight_kg * 100) if v.max_weight_kg > 0 else 0
        volume_pct = (route.total_volume / v.max_volume_cbm * 100) if v.max_volume_cbm > 0 else 0

        start_time = getattr(v, 'start_time', 480)
        if hasattr(start_time, 'hour'):
            start_time = start_time.hour * 60 + start_time.minute
        working_hours = (route.stops[-1].departure_time - start_time) / 60 if route.stops else 0
        late_count = sum(1 for s in route.stops if getattr(s, 'is_late', False))

        fixed = getattr(v, 'fixed_cost', 1500)
        variable = route.total_distance * getattr(v, 'variable_cost', 14)
        extra = max(0, len(route.stops) - getattr(v, 'capacity_drop', 7)) * getattr(v, 'extra_drop_charge', 150)
        total_cost = fixed + variable + extra

        if hasattr(route, 'orders_by_depot') and route.orders_by_depot:
            depot_ids = list(route.orders_by_depot.keys())
            depot_names = ", ".join([depots[d].name if d in depots else d for d in depot_ids])
        else:
            depot_names = default_depot.name if default_depot else "DC"

        maps_link = generate_google_maps_link(depot_lat, depot_lng, route.stops)

        data = [
            row_idx - 1, v.vehicle_id, getattr(v, 'driver_name', 'N/A'),
            getattr(v, 'vehicle_type', '4W'), depot_names, len(route.stops),
            f"{route.total_weight:.0f}", f"{weight_pct:.0f}%",
            f"{route.total_volume:.2f}", f"{volume_pct:.0f}%",
            f"{route.total_distance:.1f}", f"{working_hours:.1f}",
            late_count, f"{total_cost:,.0f}", maps_link
        ]
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx == len(data) and str(value).startswith('http'):
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")
                cell.value = "🗺️"

    # Totals
    total_row = len(routes) + 2
    active = [r for r in routes if not r.is_empty()]
    totals = [
        'TOTAL', f'{len(active)} vehicles', '', '', '',
        sum(len(r.stops) for r in active),
        f"{sum(r.total_weight for r in active):,.0f}", '',
        f"{sum(r.total_volume for r in active):.2f}", '',
        f"{sum(r.total_distance for r in active):,.1f}", '',
        sum(sum(1 for s in r.stops if getattr(s, 'is_late', False)) for r in active),
        '', ''
    ]
    for col_idx, value in enumerate(totals, 1):
        cell = ws.cell(row=total_row, column=col_idx, value=value)
        cell.border = border
        cell.font = Font(bold=True)

    ws.freeze_panes = 'A2'


def _add_unassigned_sheet(ws, unassigned, header_fill, header_font, border, priority_config):
    """Sheet Unassigned — เหมือน Part D เดิม"""
    headers = ['No', 'Customer', 'District', 'Province', 'Zone', 'Plant',
               'Priority', 'Priority_Name', 'Deadline',
               'Weight', 'Volume', 'Time_Window', 'Reason', 'Suggested_Action']

    priority_fills = {
        1: PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
        2: PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
        3: PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
        4: PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        5: PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    }

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    sorted_unassigned = sorted(unassigned, key=lambda o: getattr(o, 'priority', 4))

    for row_idx, order in enumerate(sorted_unassigned, 2):
        priority = getattr(order, 'priority', 4)
        priority_name = get_priority_name(priority, priority_config)
        deadline = get_order_deadline(order, priority_config)
        deadline_str = minutes_to_time_str(deadline)
        tw = (f"{minutes_to_time_str(getattr(order, 'time_open', 480))}-"
              f"{minutes_to_time_str(getattr(order, 'time_close', 1020))}")
        reason = getattr(order, 'unassign_reason', '') or 'ไม่ทราบสาเหตุ'
        action = "⚠️ URGENT: Add vehicle or extend hours" if priority <= 2 else "Check capacity/Add vehicle"
        data = [
            row_idx - 1, order.customer_name, getattr(order, 'district', ''),
            getattr(order, 'province', ''), getattr(order, 'zone', ''),
            getattr(order, 'plant', ''),
            priority, priority_name, deadline_str,
            f"{order.weight_kg:.1f}", f"{order.volume_cbm:.3f}", tw, reason, action
        ]
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx == 7 and priority in priority_fills:
                cell.fill = priority_fills[priority]


# =====================================================
# SECTION 5: EXPORT MATERIAL_BY_ROUTE.XLSX (3 sheets)
# =====================================================
def export_material_by_route_excel(solution, depots, raw_order_lines=None,
                                    priority_config=None):
    """
    Export Material_by_Route.xlsx — 3 sheets:
    1. Material_by_Route (Item level + สีแบ่ง Route + เส้นหนาคั่น)
    2. Summary
    3. Unassigned
    Returns: BytesIO
    """
    if priority_config is None:
        priority_config = DEFAULT_PRIORITY_CONFIG

    if raw_order_lines is None:
        raw_order_lines = pd.DataFrame()

    # ── 1) Build customer → route mapping ──
    customer_route_map = {}
    for route_idx, route in enumerate(solution.routes, 1):
        if route.is_empty():
            continue
        v = route.vehicle
        for stop in route.stops:
            cust_name = stop.order.customer_name.strip().lower()
            customer_route_map[cust_name] = {
                'Route': route_idx,
                'Driver Name': getattr(v, 'driver_name', ''),
                'Vehicle Type': getattr(v, 'vehicle_type', '4W'),
                'Priority': getattr(stop.order, 'priority', 4),
            }

    # ── 2) Map raw order lines → Route ──
    output_rows = []
    unassigned_rows = []

    if not raw_order_lines.empty:
        raw_order_lines.columns = raw_order_lines.columns.str.strip()

        for _, row in raw_order_lines.iterrows():
            cust_name = ''
            for col in ['Customer Name', 'Ship to Name']:
                if col in row.index and pd.notna(row.get(col)):
                    cust_name = str(row[col]).strip()
                    break

            route_info = customer_route_map.get(cust_name.lower(), {})

            if not route_info:
                for key, info in customer_route_map.items():
                    if cust_name.lower() in key or key in cust_name.lower():
                        route_info = info
                        break

            priority = route_info.get('Priority', 4)
            priority_name = get_priority_name(priority, priority_config) if route_info else 'N/A'

            row_dict = {
                'Plant': row.get('Plant', ''),
                'Route': route_info.get('Route', 'Unassigned'),
                'Priority': priority if route_info else 'N/A',
                'Priority_Name': priority_name,
                'DN': row.get('DN Number', ''),
                'Material Code': row.get('Mat Code', ''),
                'Delivery Qty': row.get('จำนวน', ''),
                'Sales Unit': row.get('หน่วย', ''),
                'Material Description': row.get('Item Desc', ''),
                'Barcode': row.get('Barcode', ''),
                'Weight (KG)': row.get('น้ำหนัก', 0),
                'Volume (CBM)': row.get('Vol', row.get('ปริมาตร CBM', 0)),
                'Shipto Name': cust_name,
                'Shipto District': row.get('อำเภอ', ''),
                'Shipto Province': row.get('จังหวัด', ''),
                'Driver Name': route_info.get('Driver Name', ''),
                'Vehicle Type': route_info.get('Vehicle Type', ''),
            }

            if route_info:
                output_rows.append(row_dict)
            else:
                row_dict['Route'] = 'Unassigned'
                unassigned_rows.append(row_dict)
                output_rows.append(row_dict)

    # ── 3) Sort ──
    df = pd.DataFrame(output_rows)
    if not df.empty:
        df['_sort'] = df['Route'].apply(lambda x: 9999 if x == 'Unassigned' else int(x))
        df = df.sort_values(['_sort', 'DN']).drop('_sort', axis=1)

    output = io.BytesIO()

    if not OPENPYXL_AVAILABLE or df.empty:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            if not df.empty:
                df.to_excel(writer, sheet_name='Material_by_Route', index=False)
            else:
                pd.DataFrame({'Status': ['No data']}).to_excel(
                    writer, sheet_name='Material_by_Route', index=False)
        output.seek(0)
        return output

    wb = Workbook()
    ws = wb.active
    ws.title = "Material_by_Route"

    # ── Styles ──
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    unassigned_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    thick_bottom = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thick', color='000000')
    )

    route_colors = [
        "E3F2FD", "FFF3E0", "E8F5E9", "F3E5F5",
        "FFFDE7", "E0F7FA", "FBE9E7", "F1F8E9",
    ]

    columns = list(df.columns)

    # ── Header ──
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    # ── Find last row per route ──
    last_row_of_route = {}
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        route_val = row_data.get('Route', '')
        last_row_of_route[route_val] = row_idx

    # ── Data Rows ──
    current_route = None
    route_color_idx = -1

    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        route_val = row_data.get('Route', '')

        if route_val != current_route:
            current_route = route_val
            if route_val != 'Unassigned':
                route_color_idx = (route_color_idx + 1) % len(route_colors)

        is_last = (last_row_of_route.get(route_val) == row_idx)

        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if route_val == 'Unassigned':
                cell.fill = unassigned_fill
            else:
                cell.fill = PatternFill(
                    start_color=route_colors[route_color_idx],
                    end_color=route_colors[route_color_idx],
                    fill_type="solid"
                )

            if is_last and route_val != 'Unassigned':
                cell.border = thick_bottom
            else:
                cell.border = thin_border

            if col_name in ['Plant', 'Route', 'Priority', 'Delivery Qty',
                           'Sales Unit', 'Weight (KG)', 'Volume (CBM)', 'Vehicle Type']:
                cell.alignment = Alignment(horizontal="center")

    # ── Column Widths ──
    col_widths = {
        'A': 8, 'B': 8, 'C': 8, 'D': 12, 'E': 12, 'F': 12,
        'G': 10, 'H': 10, 'I': 45, 'J': 15, 'K': 12, 'L': 12,
        'M': 40, 'N': 20, 'O': 18, 'P': 15, 'Q': 12,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = 'A2'

    # ── Sheet 2: Summary ──
    ws_summary = wb.create_sheet("Summary")
    _add_material_summary_sheet(ws_summary, df, hdr_font, hdr_fill, hdr_align,
                                 thin_border, unassigned_fill)

    # ── Sheet 3: Unassigned ──
    ws_unassigned = wb.create_sheet("Unassigned")
    _add_material_unassigned_sheet(ws_unassigned, df, hdr_font, hdr_fill,
                                    hdr_align, thin_border, unassigned_fill)

    wb.save(output)
    output.seek(0)
    return output


def _add_material_summary_sheet(ws, df, hdr_font, hdr_fill, hdr_align, border, unassigned_fill):
    """Summary sheet for Material_by_Route"""
    summary_headers = ['Route', 'Priority', 'Driver', 'Vehicle Type', 'Items',
                       'DN Count', 'Total Weight (KG)', 'Total Volume (CBM)']

    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    df_assigned = df[df['Route'] != 'Unassigned'].copy()

    if not df_assigned.empty:
        df_assigned['Weight (KG)'] = pd.to_numeric(df_assigned['Weight (KG)'], errors='coerce').fillna(0)
        df_assigned['Volume (CBM)'] = pd.to_numeric(df_assigned['Volume (CBM)'], errors='coerce').fillna(0)

        summary = df_assigned.groupby('Route').agg({
            'Priority': 'first',
            'Driver Name': 'first',
            'Vehicle Type': 'first',
            'DN': ['count', 'nunique'],
            'Weight (KG)': 'sum',
            'Volume (CBM)': 'sum'
        }).reset_index()
        summary.columns = ['Route', 'Priority', 'Driver', 'Vehicle',
                           'Items', 'DN_Unique', 'Weight', 'Volume']
        summary = summary.sort_values('Route')

        for row_idx, (_, row) in enumerate(summary.iterrows(), 2):
            data = [
                row['Route'], row['Priority'], row['Driver'], row['Vehicle'],
                row['Items'], row['DN_Unique'],
                round(row['Weight'], 2), round(row['Volume'], 4)
            ]
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if col_idx in [1, 2, 4, 5, 6]:
                    cell.alignment = Alignment(horizontal="center")

        total_row = len(summary) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=summary['Items'].sum())
        ws.cell(row=total_row, column=6, value=summary['DN_Unique'].sum())
        ws.cell(row=total_row, column=7, value=round(summary['Weight'].sum(), 2))
        ws.cell(row=total_row, column=8, value=round(summary['Volume'].sum(), 4))
        for col_idx in range(1, 9):
            ws.cell(row=total_row, column=col_idx).border = border
            ws.cell(row=total_row, column=col_idx).font = Font(bold=True)

    summary_widths = {'A': 10, 'B': 10, 'C': 15, 'D': 12,
                      'E': 10, 'F': 12, 'G': 18, 'H': 18}
    for col, width in summary_widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = 'A2'


def _add_material_unassigned_sheet(ws, df, hdr_font, hdr_fill, hdr_align,
                                    border, unassigned_fill):
    """Unassigned sheet for Material_by_Route"""
    df_unassigned = df[df['Route'] == 'Unassigned'].copy()

    headers = ['Shipto Name', 'Plant', 'DN', 'Material Code',
               'Material Description', 'Delivery Qty', 'Sales Unit',
               'Weight (KG)', 'Volume (CBM)', 'Shipto District', 'Shipto Province']

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    if df_unassigned.empty:
        ws.cell(row=2, column=1, value="✅ ทุกรายการจัดเข้า Route ได้หมด")
        return

    for row_idx, (_, row_data) in enumerate(df_unassigned.iterrows(), 2):
        for col_idx, col_name in enumerate(headers, 1):
            value = row_data.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = unassigned_fill
            cell.border = border

    col_widths = {'A': 40, 'B': 8, 'C': 12, 'D': 12, 'E': 45,
                  'F': 10, 'G': 10, 'H': 12, 'I': 12, 'J': 20, 'K': 18}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = 'A2'


# =====================================================
# SECTION 6: ZIP DOWNLOAD (2 files)
# =====================================================
def create_download_zip(solution, data, priority_config=None):
    """
    สร้าง zip ที่มี 2 ไฟล์:
    1. Route_Detail.xlsx
    2. Material_by_Route.xlsx
    Returns: BytesIO (zip)
    """
    if priority_config is None:
        if hasattr(data, 'priority_config'):
            priority_config = data.priority_config
        elif isinstance(data, dict):
            priority_config = data.get('priority_config', DEFAULT_PRIORITY_CONFIG)
        else:
            priority_config = DEFAULT_PRIORITY_CONFIG

    depots = {}
    raw_order_lines = pd.DataFrame()
    if hasattr(data, 'depots'):
        depots = data.depots
        raw_order_lines = data.raw_order_lines if data.raw_order_lines is not None else pd.DataFrame()
    elif isinstance(data, dict):
        depots = data.get('depots', {})
        raw_order_lines = data.get('raw_order_lines', pd.DataFrame())

    # Generate both files
    route_detail_bytes = export_route_detail_excel(solution, depots, priority_config)
    material_bytes = export_material_by_route_excel(
        solution, depots, raw_order_lines, priority_config
    )

    # Create ZIP
    zip_buffer = io.BytesIO()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"Route_Detail_{timestamp}.xlsx", route_detail_bytes.getvalue())
        zf.writestr(f"Material_by_Route_{timestamp}.xlsx", material_bytes.getvalue())

    zip_buffer.seek(0)
    return zip_buffer


# =====================================================
# SECTION 7: MATPLOTLIB CHARTS
# =====================================================
def generate_utilization_chart(summary_report):
    """
    กราฟ Weight % vs Volume % — matplotlib
    เหมือนรูปเดิมเป๊ะ: steelblue + coral + เส้นประ 80%
    Returns: matplotlib Figure
    """
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    route_summaries = summary_report['route_summaries']
    if not route_summaries:
        return None

    vehicle_ids = [rs['vehicle_id'] for rs in route_summaries]
    weight_utils = [rs['weight_util'] for rs in route_summaries]
    volume_utils = [rs['volume_util'] for rs in route_summaries]

    fig, ax = plt.subplots(figsize=(14, 6))

    x = np.arange(len(vehicle_ids))
    width = 0.35

    ax.bar(x - width/2, weight_utils, width, label='Weight %', color='steelblue')
    ax.bar(x + width/2, volume_utils, width, label='Volume %', color='coral')

    ax.set_title('Solution Statistics', fontsize=12, fontweight='bold')
    ax.set_ylabel('Utilization (%)')
    ax.set_xlabel('Vehicle')
    ax.set_xticks(x)
    ax.set_xticklabels(vehicle_ids, rotation=45, ha='right')
    ax.legend()
    ax.axhline(y=80, color='green', linestyle='--', alpha=0.5, label='80% target')
    ax.set_ylim(0, 110)
    ax.grid(True, axis='y', alpha=0.3)

    plt.tight_layout()
    return fig


def generate_distance_chart(summary_report):
    """
    กราฟ Pickup vs Delivery Distance — matplotlib
    Returns: matplotlib Figure
    """
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    route_summaries = summary_report['route_summaries']
    if not route_summaries:
        return None

    labels = [f"R{rs['route_number']}" for rs in route_summaries]
    pickup = [rs['pickup_distance'] for rs in route_summaries]
    delivery = [rs['delivery_distance'] for rs in route_summaries]

    fig, ax = plt.subplots(figsize=(14, 6))

    x = np.arange(len(labels))
    width = 0.35

    ax.bar(x - width/2, pickup, width, label='Pickup Distance', color='#3498db')
    ax.bar(x + width/2, delivery, width, label='Delivery Distance', color='#e74c3c')

    ax.set_title('Distance Breakdown by Route', fontsize=12, fontweight='bold')
    ax.set_ylabel('Distance (km)')
    ax.set_xlabel('Route')
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=45, ha='right')
    ax.legend()
    ax.grid(True, axis='y', alpha=0.3)

    plt.tight_layout()
    return fig


def generate_convergence_chart(operator_stats):
    """
    กราฟ Cost Convergence — matplotlib
    Returns: matplotlib Figure
    """
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    cost_history = operator_stats.get('history', {}).get('cost', [])
    best_history = operator_stats.get('history', {}).get('best_cost', [])

    if not best_history:
        return None

    fig, ax = plt.subplots(figsize=(14, 5))

    iterations = range(1, len(best_history) + 1)
    ax.plot(iterations, cost_history, alpha=0.3, color='gray', linewidth=0.5, label='Current')
    ax.plot(iterations, best_history, color='#e74c3c', linewidth=1.5, label='Best')

    best_iter = operator_stats.get('best_iteration', 0)
    if best_iter > 0 and best_iter <= len(best_history):
        ax.axvline(x=best_iter, color='green', linestyle='--', alpha=0.5, label=f'Best at #{best_iter}')

    ax.set_title('ALNS Cost Convergence', fontsize=12, fontweight='bold')
    ax.set_xlabel('Iteration')
    ax.set_ylabel('Cost')
    ax.legend()
    ax.grid(True, alpha=0.3)

    plt.tight_layout()
    return fig


def generate_temperature_chart(operator_stats):
    """
    กราฟ Temperature — matplotlib
    Returns: matplotlib Figure
    """
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    temp_history = operator_stats.get('history', {}).get('temperature', [])

    if not temp_history:
        return None

    fig, ax = plt.subplots(figsize=(14, 4))

    iterations = range(1, len(temp_history) + 1)
    ax.plot(iterations, temp_history, color='#f39c12', linewidth=1.0)

    # Mark reheat points
    reheat_count = operator_stats.get('reheat_count', 0)
    if reheat_count > 0:
        # Find jumps in temperature (reheat points)
        for i in range(1, len(temp_history)):
            if temp_history[i] > temp_history[i-1] * 1.5:
                ax.axvline(x=i, color='red', linestyle=':', alpha=0.7)
                ax.annotate(f'🔥 Reheat', xy=(i, temp_history[i]),
                           fontsize=8, color='red')

    ax.set_title(f'Temperature Schedule (Reheats: {reheat_count})',
                 fontsize=12, fontweight='bold')
    ax.set_xlabel('Iteration')
    ax.set_ylabel('Temperature')
    ax.grid(True, alpha=0.3)
    ax.set_yscale('log')

    plt.tight_layout()
    return fig


# =====================================================
# SECTION 8: MAP GENERATION (Folium) — v3.3 แก้หน้าขาว
# =====================================================
def generate_map(solution, data, summary_report=None):
    """
    สร้าง Folium Map
    v3.3: แก้ปัญหาหน้าขาว
    - ป้องกัน None lat/lng
    - ป้องกัน crash เงียบ
    - return Folium Map object (ไม่ใช่ HTML string)
    - ให้ app.py เป็นคนแปลง HTML เอง
    Returns: folium.Map object หรือ None
    """
    try:
        import folium
    except ImportError:
        return None

    # ── ดึง depots ──
    depots = {}
    if data is not None:
        if hasattr(data, 'depots'):
            depots = data.depots
        elif isinstance(data, dict):
            depots = data.get('depots', {})

    # ── รวบรวมพิกัดทั้งหมด (กรอง None/0) ──
    all_lats = []
    all_lngs = []

    for d in depots.values():
        lat = getattr(d, 'lat', 0) or 0
        lng = getattr(d, 'lng', 0) or 0
        if lat != 0 and lng != 0:
            all_lats.append(float(lat))
            all_lngs.append(float(lng))

    for route in solution.routes:
        if route.is_empty():
            continue
        for stop in route.stops:
            order = stop.order
            lat = getattr(order, 'lat', 0) or 0
            lng = getattr(order, 'lng', 0) or 0
            if lat != 0 and lng != 0:
                all_lats.append(float(lat))
                all_lngs.append(float(lng))

    if not all_lats:
        center_lat, center_lng = 14.0, 100.5
    else:
        center_lat = sum(all_lats) / len(all_lats)
        center_lng = sum(all_lngs) / len(all_lngs)

    # ── สร้างแผนที่ ──
    m = folium.Map(
        location=[center_lat, center_lng],
        zoom_start=10,
        tiles='OpenStreetMap'
    )

    # ── Depot Markers ──
    for depot_id, depot in depots.items():
        lat = getattr(depot, 'lat', 0) or 0
        lng = getattr(depot, 'lng', 0) or 0
        if lat == 0 or lng == 0:
            continue

        icon_color = DEPOT_ICON_COLORS.get(depot_id, 'gray')
        depot_name = getattr(depot, 'name', str(depot_id))

        try:
            folium.Marker(
                location=[float(lat), float(lng)],
                popup=folium.Popup(
                    f"<b>🏭 {depot_name}</b><br>ID: {depot_id}",
                    max_width=250
                ),
                tooltip=f"Depot: {depot_name}",
                icon=folium.Icon(
                    color=icon_color, icon='industry', prefix='fa'
                )
            ).add_to(m)
        except Exception:
            pass

    # ── Route Lines + Stop Markers ──
    active_routes = [
        r for r in solution.routes if not r.is_empty()
    ]

    for route_idx, route in enumerate(active_routes):
        color = ROUTE_COLORS[route_idx % len(ROUTE_COLORS)]
        vehicle = route.vehicle
        vtype = getattr(vehicle, 'vehicle_type', '4W')
        vid = getattr(vehicle, 'vehicle_id', '?')
        n_stops = len(route.stops)

        route_group = folium.FeatureGroup(
            name=f"R{route_idx+1}: {vid} ({vtype}) "
                 f"- {n_stops} stops"
        )

        # ── PICKUP PATH (depot → depot) ──
        try:
            pickup_seq = getattr(route, 'pickup_sequence', [])
            if pickup_seq and len(pickup_seq) > 0:
                pickup_coords = []
                for depot in pickup_seq:
                    dlat = getattr(depot, 'lat', 0) or 0
                    dlng = getattr(depot, 'lng', 0) or 0
                    if dlat != 0 and dlng != 0:
                        pickup_coords.append(
                            [float(dlat), float(dlng)]
                        )

                # วาดเส้น pickup (dash)
                if len(pickup_coords) >= 2:
                    folium.PolyLine(
                        pickup_coords,
                        weight=4, color=color,
                        opacity=0.9, dash_array='10, 5',
                        tooltip=(
                            f"Pickup: "
                            f"{' → '.join(getattr(d, 'name', '?') for d in pickup_seq)}"
                        )
                    ).add_to(route_group)

                # Pickup markers
                for pi, depot in enumerate(pickup_seq):
                    dlat = getattr(depot, 'lat', 0) or 0
                    dlng = getattr(depot, 'lng', 0) or 0
                    if dlat == 0 or dlng == 0:
                        continue
                    dname = getattr(depot, 'name', '?')

                    folium.Marker(
                        location=[float(dlat), float(dlng)],
                        icon=folium.DivIcon(
                            html=(
                                f'<div style="font-size:11px;'
                                f'font-weight:bold;color:white;'
                                f'background:{color};'
                                f'border:2px solid white;'
                                f'border-radius:4px;'
                                f'width:24px;height:18px;'
                                f'text-align:center;'
                                f'line-height:18px;">'
                                f'P{pi+1}</div>'
                            ),
                            icon_size=(24, 18),
                            icon_anchor=(12, 9),
                        ),
                        tooltip=f"Pickup #{pi+1}: {dname}"
                    ).add_to(route_group)
        except Exception:
            pass

        # ── DELIVERY PATH ──
        delivery_coords = []

        # Start from last pickup depot
        try:
            pickup_seq = getattr(route, 'pickup_sequence', [])
            if pickup_seq:
                last_d = pickup_seq[-1]
                dlat = getattr(last_d, 'lat', 0) or 0
                dlng = getattr(last_d, 'lng', 0) or 0
                if dlat != 0 and dlng != 0:
                    delivery_coords.append(
                        [float(dlat), float(dlng)]
                    )
        except Exception:
            pass

        # ── Stop Markers ──
        for stop_idx, stop in enumerate(route.stops):
            try:
                order = stop.order
                lat = getattr(order, 'lat', 0) or 0
                lng = getattr(order, 'lng', 0) or 0

                if lat == 0 or lng == 0:
                    continue

                lat = float(lat)
                lng = float(lng)
                delivery_coords.append([lat, lng])

                # ── ดึงข้อมูลอย่างปลอดภัย ──
                cust_name = getattr(
                    order, 'customer_name', '?'
                )[:40]
                is_late = bool(
                    _safe_get(stop, 'is_late', False)
                )
                arrival = _safe_get(stop, 'arrival_time', 0)
                weight = getattr(order, 'weight_kg', 0) or 0
                volume = getattr(order, 'volume_cbm', 0) or 0
                plant = getattr(order, 'plant', '')
                dist_prev = _safe_get(
                    stop, 'distance_from_prev', 0
                )
                t_open = getattr(order, 'time_open', 480)
                t_close = getattr(order, 'time_close', 1020)

                late_str = '⚠️ LATE' if is_late else '✅'

                popup_html = (
                    f'<div style="width:240px;'
                    f'font-family:sans-serif;'
                    f'font-size:12px;">'
                    f'<b>Stop {stop_idx+1}: '
                    f'{cust_name}</b><br>'
                    f'{late_str}<br>'
                    f'📦 {weight:.0f} kg / '
                    f'{volume:.2f} CBM<br>'
                    f'⏰ Arrival: '
                    f'{minutes_to_time_str(arrival)}<br>'
                    f'🕐 TW: '
                    f'{minutes_to_time_str(t_open)}-'
                    f'{minutes_to_time_str(t_close)}<br>'
                    f'🏭 Plant: {plant}<br>'
                    f'📏 Dist: {dist_prev:.1f} km'
                    f'</div>'
                )

                # Circle marker
                folium.CircleMarker(
                    location=[lat, lng],
                    radius=8,
                    popup=folium.Popup(
                        popup_html, max_width=280
                    ),
                    tooltip=(
                        f"Stop {stop_idx+1}: {cust_name}"
                    ),
                    color='red' if is_late else color,
                    fill=True,
                    fillColor=color,
                    fillOpacity=0.7,
                    weight=2,
                ).add_to(route_group)

                # Number label
                folium.Marker(
                    location=[lat, lng],
                    icon=folium.DivIcon(
                        html=(
                            f'<div style="font-size:10px;'
                            f'font-weight:bold;color:white;'
                            f'background:{color};'
                            f'border-radius:50%;'
                            f'width:20px;height:20px;'
                            f'text-align:center;'
                            f'line-height:20px;">'
                            f'{stop_idx+1}</div>'
                        ),
                        icon_size=(20, 20),
                        icon_anchor=(10, 10),
                    )
                ).add_to(route_group)

            except Exception:
                continue

        # วาดเส้น delivery
        if len(delivery_coords) >= 2:
            try:
                total_dist = _safe_get(
                    route, 'total_distance', 0
                )
                folium.PolyLine(
                    delivery_coords,
                    weight=3, color=color, opacity=0.7,
                    tooltip=(
                        f"R{route_idx+1} {vid}: "
                        f"{total_dist:.1f} km"
                    )
                ).add_to(route_group)
            except Exception:
                pass

        route_group.add_to(m)

    # ── Unassigned Markers ──
    if solution.unassigned:
        unassigned_group = folium.FeatureGroup(
            name=f"❌ Unassigned ({len(solution.unassigned)})"
        )
        for order in solution.unassigned:
            try:
                lat = getattr(order, 'lat', 0) or 0
                lng = getattr(order, 'lng', 0) or 0
                if lat == 0 or lng == 0:
                    continue
                cust = getattr(
                    order, 'customer_name', '?'
                )[:30]
                weight = getattr(order, 'weight_kg', 0)
                plant = getattr(order, 'plant', '')

                folium.CircleMarker(
                    location=[float(lat), float(lng)],
                    radius=6,
                    popup=(
                        f"❌ {cust}<br>"
                        f"Plant: {plant}<br>"
                        f"Weight: {weight:.0f} kg"
                    ),
                    tooltip=f"Unassigned: {cust}",
                    color='black', fill=True,
                    fillColor='black', fillOpacity=0.5,
                ).add_to(unassigned_group)
            except Exception:
                continue
        unassigned_group.add_to(m)

    # ── Layer Control ──
    folium.LayerControl(collapsed=False).add_to(m)

    # ── Legend ──
    try:
        total_orders = sum(
            len(r.stops) for r in active_routes
        )
        total_dist = sum(
            _safe_get(r, 'total_distance', 0)
            for r in active_routes
        )
        un_count = len(solution.unassigned) if solution.unassigned else 0

        legend_html = f"""
        <div style="position:fixed; bottom:30px; left:30px;
                    z-index:1000; background:white;
                    padding:12px; border-radius:8px;
                    border:2px solid #333;
                    font-size:12px;
                    font-family:sans-serif;
                    box-shadow:2px 2px 6px rgba(0,0,0,0.3);">
            <b>🗺️ VRP Solution</b><br>
            🚛 Routes: {len(active_routes)}<br>
            📦 Orders: {total_orders}<br>
            📏 Distance: {total_dist:,.1f} km<br>
            ❌ Unassigned: {un_count}
        </div>
        """
        m.get_root().html.add_child(folium.Element(legend_html))
    except Exception:
        pass

    return m  # ← return Map object ไม่ใช่ HTML string


# =====================================================
# SECTION 9: CONVERGENCE DATA (for ALNS tab)
# =====================================================
def generate_convergence_data(operator_stats):
    """สร้างข้อมูล Convergence"""
    return {
        'cost_history': operator_stats.get('history', {}).get('cost', []),
        'best_cost_history': operator_stats.get('history', {}).get('best_cost', []),
        'temperature_history': operator_stats.get('history', {}).get('temperature', []),
        'iterations': operator_stats.get('iterations', 0),
        'best_iteration': operator_stats.get('best_iteration', 0),
        'reheat_count': operator_stats.get('reheat_count', 0),
        'elapsed': operator_stats.get('elapsed', 0),
        'destroy_operators': operator_stats.get('destroy', []),
        'repair_operators': operator_stats.get('repair', []),
    }